'''
Desc: 操作文件工具函数
File: /file_op.py
Project: utils
File Created: Sunday, 9th May 2021 5:04:11 pm
Author: luxuemin2108@gmail.com
-----
Copyright (c) 2021 Camel Lu
'''
import json
import os
import time

import pandas as pd
from openpyxl import load_workbook

from .index import get_last_quarter_str


# 写json文件
def write_fund_json_data(data, filename, file_dir=None):
    import json
    if not file_dir:
        cur_date = time.strftime("%Y-%m-%d", time.localtime(time.time()))
        file_dir = os.getcwd() + '/output/json/ai_fund/' + cur_date + '/'
    if not os.path.exists(file_dir):
        os.makedirs(file_dir)
        print("目录新建成功：%s" % file_dir)
    with open(file_dir + filename, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
        f.close()

def read_dir_all_file(path):
    return os.listdir(path)

def update_xlsx_file(path, df_data, sheet_name):
    try:
        if os.path.exists(path):
            writer = pd.ExcelWriter(path, engine='openpyxl')
            book = load_workbook(path)
            # 表名重复，删掉，重写
            if sheet_name in book.sheetnames:
                del book[sheet_name]
            if len(book.sheetnames) == 0:
                df_data.to_excel(
                    path, sheet_name=sheet_name)
                return
            else:
                writer.book = book
            df_data.to_excel(
                    writer, sheet_name=sheet_name)

            writer.save()
            writer.close()
        else:
            df_data.to_excel(
                path, sheet_name=sheet_name)
    except BaseException:
        print("path", path)
        raise BaseException('更新excel失败')


def update_xlsx_file_with_sorted(path, df_data, sheet_name, sorted_sheetnames = []):
    try:
        if os.path.exists(path):
            writer = pd.ExcelWriter(path, engine='openpyxl')
            workbook = load_workbook(path)
            writer.book = workbook
            writer.sheets = {ws.title:ws for ws in workbook.worksheets}
            for sheet_item in sorted_sheetnames:
                del workbook[sheet_item]
            df_data.to_excel(
                    writer,  sheet_name=sheet_name)
            workbook = writer.book
            for worksheet in sorted_sheetnames:
                workbook._add_sheet(writer.sheets.get(worksheet))
            writer.book = workbook

            writer.save()
            writer.close()
        else:
            df_data.to_excel(
                path, sheet_name=sheet_name)
    except BaseException:
        print("path", path)
        raise BaseException('更新excel失败')

def update_xlsx_file_with_insert(path, df_data, sheet_name, index = 0):
    try:
        if os.path.exists(path):
            writer = pd.ExcelWriter(path, engine='openpyxl')
            workbook = load_workbook(path)
            if sheet_name in workbook.sheetnames:
                del workbook[sheet_name]
            writer.book = workbook
            df_data.to_excel(
                    writer,  sheet_name=sheet_name)
            workbook = writer.book
            writer.sheets = {ws.title:ws for ws in workbook.worksheets}
            # workbook.remove(sheet_name)
            del workbook[sheet_name]
            
            workbook._add_sheet(writer.sheets.get(sheet_name), index)
            writer.book = workbook

            writer.save()
            writer.close()
        else:
            df_data.to_excel(
                path, sheet_name=sheet_name)
    except BaseException:
        print("path", path)
        raise BaseException('更新excel失败')

def read_error_code_from_json():
    quarter_index = get_last_quarter_str()
    filename = 'error_funds_' + quarter_index + '.json'
    file_dir = './output/json/'
    error_funds_with_page = []
    error_funds_with_unmatch = []
    error_funds_with_found_date = []
    if os.path.exists(file_dir + filename):
        with open(file_dir + filename) as json_file:
            my_data = json.load(json_file)
            error_funds_with_page = my_data.get('error_funds_with_page')
            error_funds_with_found_date = my_data.get('error_funds_with_found_date')
    return {
        "file_dir": file_dir,
        "filename": filename,
        'error_funds_with_unmatch': error_funds_with_unmatch,
        'error_funds_with_page': error_funds_with_page,
        'error_funds_with_found_date': error_funds_with_found_date
    }
