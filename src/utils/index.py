
import time
import datetime
import os
import numpy as np
import requests
from PIL import Image
from skimage import io
from sewar.full_ref import uqi, sam
import re
from threading import Thread, Lock

import pandas as pd
from openpyxl import load_workbook

requests.adapters.DEFAULT_RETRIES = 10 # 增加重连次数
s = requests.session()
s.keep_alive = False # 关闭多余连接

dir = os.getcwd() + '/src/'

img_dir = dir + 'img/'
samples_dir = dir + 'assets/samples/'

def use_sewar_get_star_level(img_path):
    sample_imgs = os.listdir(samples_dir)
    img1 = io.imread(fname=img_path)
    for filename in sample_imgs:
        level = filename[-5:-4]
        img_path_2 = samples_dir + filename
        img2 = io.imread(fname=img_path_2)
        res_uqi = uqi(img1, img2)
        res_sam = sam(img1, img2)

        if res_uqi > 0.98 and res_sam < 0.11:
            # res_level = level2
            return level
    print('img_path', img_path)
    raise "img_path 图片比较失败"
def lock_process(func):
    lock = Lock()

    def wrapper(self, *args):
        lock.acquire()
        result = func(self, *args)
        lock.release()
        return result
    return wrapper


def debug(func):
    def wrapper(self, *args):  # 指定一毛一样的参数
        print("[DEBUG]: enter {}()".format(func.__name__))
        return func(self, *args)
    return wrapper  # 返回包装过函数


def get_star_count_with_sewar(fund_code, img_ele):
    picture_time = time.strftime(
        "%Y-%m-%d-%H_%M_%S", time.localtime(time.time()))
    directory_time = time.strftime("%Y-%m-%d", time.localtime(time.time()))
    file_dir = os.getcwd() + '/star-record/' + directory_time + '/'
    try:
        if not os.path.exists(file_dir):
            os.makedirs(file_dir)
            print("目录新建成功：%s" % file_dir)
    except BaseException as msg:
        print("新建目录失败：%s" % msg)
    
    code_path = './star-record/' + directory_time + '/' + picture_time + '_' + fund_code + '_' + '_code.png'
    is_success = img_ele.screenshot(code_path)
    time.sleep(2)
    if is_success:
        return use_sewar_get_star_level(code_path)
    else:
        raise "截图失败"
    

def get_star_count_with_np(morning_star_url):
    module_path = os.getcwd() + '/src'
    temp_star_url = module_path + '/assets/star/tmp.gif'
    try:
        r = requests.get(morning_star_url)
    except BaseException:
        raise BaseException('请求失败')
    with open(temp_star_url, "wb") as f:
        f.write(r.content)
    f.close()
    path = module_path + '/assets/star/star'

    try:
        for i in range(6):
            p1 = np.array(Image.open(path + str(i) + '.gif'))
            p2 = np.array(Image.open(temp_star_url))
            if (p1 == p2).all():
                return i
    except BaseException:
        raise BaseException('识别失败')

def get_star_count(morning_star_url, fund_code, img_ele=None):
    # path = './assets/star/star'
    try:
        return get_star_count_with_sewar(fund_code, img_ele)
    except BaseException:
        print("BaseException", BaseException)
        print('图片相似度比较失败')
    return get_star_count_with_np(morning_star_url)
    


def parse_csv(datafile):
    data = []
    with open(datafile, "r") as f:
        header = f.readline().split(",")  # 获取表头
        counter = 0
        for line in f:
            if counter == 10:
                break
            fields = line.split(",")
            entry = {}
            for i, value in enumerate(fields):
                entry[header[i].strip()] = value.strip()  # 用strip方法去除空白
            data.append(entry)
            counter += 1

    return data


def get_quarter_index(input_date):
    year = time.strftime("%Y", time.localtime())
    boundary_date_list = ['03-31', '06-30', '09-30', '12-31']

    input_date_strptime = datetime.datetime.strptime(
        year + '-' + input_date, '%Y-%m-%d')
    index = 1
    for idx in range(len(boundary_date_list)):
        join_date = year + '-' + boundary_date_list[idx]
        season_date_strptime = datetime.datetime.strptime(
            join_date, '%Y-%m-%d')
        if input_date_strptime <= season_date_strptime:
            index = idx + 1
            break
    return index


def get_last_quarter_str(last_index=1):
    last_quarter_time = time.localtime(
        time.time() - last_index * 3 * 30 * 24 * 3600)
    year = time.strftime("%Y", last_quarter_time)
    date = time.strftime("%m-%d", last_quarter_time)
    index = get_quarter_index(date)
    quarter_index_str = year + '-Q' + str(index)
    return quarter_index_str


def get_quarter_date(quarter_index_str):
    year = quarter_index_str.split('-')[0]
    boundary_date_list = ['03-31', '06-30', '09-30', '12-31']
    quarter_index = quarter_index_str.split('-')[1][1:]
    return year + '-' + boundary_date_list[int(quarter_index) - 1]


def fisrt_match_condition_from_list(list, code):
    for item in list:
        stock_code = item.split('-', 1)[0]
        is_exist = code == stock_code
        if is_exist:
            return item


def dict_list_to_list_list(dict_list, key_sort_list):
    tuple_list = []
    for item in dict_list:
        temp_list = []
        for key in key_sort_list:
            temp_list.append(item.get(key))
        tuple_list.append(temp_list)
    return tuple_list


def find_from_list_of_dict(dict_list, match_key, value):
    res = None
    for sub in dict_list:
        if sub[match_key] == value:
            res = sub
            break
    return res


def get_stock_market(stock_code):
    if bool(re.search("^\d{5}$", stock_code)):
        return '港股'
    elif bool(re.search("^\d{6}$", stock_code)) and bool(re.search(
            "^(00(0|1|2|3)\d{3})|(30(0|1)\d{3})|(60(0|1|2|3|5)\d{3})|68(8|9)\d{3}$", stock_code)):
        return 'A股'
    else:
        return '其他'


def update_xlsx_file(path, df_data, sheet_name, sorted_worksheets = []):
    try:
        if os.path.exists(path):
            writer = pd.ExcelWriter(path, engine='openpyxl')
            book = load_workbook(path)
            # 表名重复，删掉，重写
            if sheet_name in book.sheetnames:
                del book[sheet_name]
            if len(book.sheetnames) == 0:
                df_data.to_excel(
                    path, sheet_name=sheet_name)
                return
            else:
                writer.book = book
            df_data.to_excel(
                    writer, sheet_name=sheet_name)

            writer.save()
            writer.close()
        else:
            df_data.to_excel(
                path, sheet_name=sheet_name)
    except BaseException:
        print("path", path)
        raise BaseException('更新excel失败')


def update_xlsx_file_with_sorted(path, df_data, sheet_name, sorted_sheetnames = []):
    try:
        if os.path.exists(path):
            writer = pd.ExcelWriter(path, engine='openpyxl')
            workbook = load_workbook(path)
            writer.book = workbook
            writer.sheets = {ws.title:ws for ws in workbook.worksheets}
            for sheet_item in sorted_sheetnames:
                del workbook[sheet_item]
            df_data.to_excel(
                    writer,  sheet_name=sheet_name)
            workbook = writer.book
            for worksheet in sorted_sheetnames:
                workbook._add_sheet(writer.sheets.get(worksheet))
            writer.book = workbook

            writer.save()
            writer.close()
        else:
            df_data.to_excel(
                path, sheet_name=sheet_name)
    except BaseException:
        print("path", path)
        raise BaseException('更新excel失败')

def update_xlsx_file_with_insert(path, df_data, sheet_name, index = 0):
    try:
        if os.path.exists(path):
            writer = pd.ExcelWriter(path, engine='openpyxl')
            workbook = load_workbook(path)
            if sheet_name in workbook.sheetnames:
                del workbook[sheet_name]
            writer.book = workbook
            df_data.to_excel(
                    writer,  sheet_name=sheet_name)
            workbook = writer.book
            writer.sheets = {ws.title:ws for ws in workbook.worksheets}
            # workbook.remove(sheet_name)
            del workbook[sheet_name]
            
            workbook._add_sheet(writer.sheets.get(sheet_name), index)
            writer.book = workbook

            writer.save()
            writer.close()
        else:
            df_data.to_excel(
                path, sheet_name=sheet_name)
    except BaseException:
        print("path", path)
        raise BaseException('更新excel失败')

def bootstrap_thread(target_fn, total, thread_count=2):
    threaders = []
    start_time = time.time()
    # 如果少于10个，只开一个线程
    if total < 10:
        thread_count = 1
    step_num = total / thread_count
    for i in range(thread_count):
        # start = steps[i]['start']
        # end = steps[i]['end']
        start = i * step_num
        end = (i + 1) * step_num
        t = Thread(target=target_fn, args=(int(start), int(end)))
        t.setDaemon(True)
        threaders.append(t)
        t.start()
    for threader in threaders:
        threader.join()
    end_time = time.time()
    print(total, 'run time is %s' % (end_time - start_time))
