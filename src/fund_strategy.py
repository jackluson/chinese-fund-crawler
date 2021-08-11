'''
Desc: 一些基金策略，方案
File: /agree_strategy.py
Project: src
File Created: Monday, 24th May 2021 11:50:25 am
Author: luxuemin2108@gmail.com
-----
Copyright (c) 2021 Camel Lu
'''

import os
from sql_model.fund_query import FundQuery
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
from utils.index import get_last_quarter_str
from pprint import pprint


def output_high_score_funds(each_query, quarter_index=None):
    """
    输出高分基金
    """
    if quarter_index == None:
        quarter_index = get_last_quarter_str()
    print("quarter_index", quarter_index)
    high_score_funds = each_query.select_high_score_funds(
        quarter_index=quarter_index)
    columns_bk = ['代码', '名称', '季度', '总资产', '现任基金经理管理起始时间', '投资风格', '三月最大回撤', '六月最大回撤', '夏普比率', '阿尔法系数', '贝塔系数',
                  'R平方', '标准差', '风险系数', '两年风险评级', '三年风险评级', '五年风险评级', '五年晨星评级', '三年晨星评级', '股票仓位', '十大持股仓位']
    columns = ['代码', '名称', '投资风格', '基金经理', '现任经理管理起始时间', '成立时间', '三年晨星评级', '五年晨星评级', '夏普比率', '股票仓位', '十大持股仓位',
               '两年风险评级', '三年风险评级', '五年风险评级', '阿尔法系数', '贝塔系数', '标准差', '总资产', '数据更新时间']
    df_high_score_funds = pd.DataFrame(high_score_funds, columns=columns)

    pprint(df_high_score_funds)
    # df_high_score_funds.to_excel(
    #     './output/xlsx/high-score-funds_log.xlsx', sheet_name=quarter_index)
    # with pd.ExcelWriter('./output/xlsx/high-score-funds_log.xlsx', engine='openpyxl') as writer:
    #     df_high_score_funds.to_excel(writer, sheet_name=quarter_index)
    # df_high_score_funds.to_excel(writer, sheet_name=quarter_index)
    # df2.to_excel(writer, sheet_name='Sheet2')
    path = './output/xlsx/high-score-funds_log.xlsx'
    if os.path.exists(path):
        writer = pd.ExcelWriter(path, engine='openpyxl')
        book = load_workbook(path)
        writer.book = book
        # 表名重复，删掉，重写
        if quarter_index in book.sheetnames:
            del book[quarter_index]

        if len(book.sheetnames) == 0:
            df_high_score_funds.to_excel(
                path, sheet_name=quarter_index)
        else:
            writer.book = book
            df_high_score_funds.to_excel(
                writer, sheet_name=quarter_index)
        writer.save()
        writer.close()
    else:
        df_high_score_funds.to_excel(
            path, sheet_name=quarter_index)


if __name__ == '__main__':
    each_query = FundQuery()
    output_high_score_funds(each_query)
